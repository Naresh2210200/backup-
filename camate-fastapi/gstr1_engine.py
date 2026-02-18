"""
GSTR1 Excel Generation Engine
==============================
Backend implementation of the frontend GSTR1 CSV processor logic.
Supports Standard & Tally formats, uses the provided Excel template.
"""
import io
import csv
import logging
import re
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")

logger = logging.getLogger(__name__)

# ─── Mappings (Using Standard for now as requested) ───────────────────────────
STANDARD_MAPPINGS: Dict[str, Dict[str, str]] = {
    'b2b': {
        'GSTIN/UIN': 'GSTIN/UIN of Recipient',
        'Invoice No': 'Invoice Number',
        'Date of Invoice': 'Invoice date',
        'Invoice Value': 'Invoice Value',
        'GST%': 'Rate',
        'Taxable Value': 'Taxable Value',
        'CESS': 'Cess Amount',
        'Place Of Supply': 'Place Of Supply',
        'RCM Applicable': 'Reverse Charge',
        'Invoice Type': 'Invoice Type',
        'E-Commerce GSTIN': 'E-Commerce GSTIN'
    },
    'b2cl': {
        'Invoice No': 'Invoice Number',
        'Date of Invoice': 'Invoice date',
        'Invoice Value': 'Invoice Value',
        'Place Of Supply': 'Place Of Supply',
        'GST%': 'Rate',
        'Taxable Value': 'Taxable Value',
        'CESS': 'Cess Amount',
        'E-Commerce GSTIN': 'E-Commerce GSTIN'
    },
    'b2cs': {
        'Type': 'Type',
        'Place Of Supply': 'Place Of Supply',
        'GST%': 'Rate',
        'Taxable Value': 'Taxable Value',
        'CESS': 'Cess Amount',
        'E-Commerce GSTIN': 'E-Commerce GSTIN'
    },
    'export': {
        'Export Type': 'Export Type',
        'Invoice No': 'Invoice Number',
        'Date of Invoice': 'Invoice date',
        'Invoice Value': 'Invoice Value',
        'Port Code': 'Port Code',
        'Shipping Bill No': 'Shipping Bill Number',
        'Shipping Bill Date': 'Shipping Bill Date',
        'GST%': 'Rate',
        'Taxable Value': 'Taxable Value'
    },
    'Nil_exempt_NonGST': {
        'Description': 'Description',
        'Nil Rated Supplies': 'Nil Rated Supplies',
        'Exempted(other than nil rated/non GST supply)': 'Exempted (other than nil rated/non GST supply)',
        'Non-GST Supplies': 'Non-GST supplies'
    },
    'cdnr': {
        'GSTIN/UIN': 'GSTIN/UIN of Recipient',
        'Dr./ Cr. No.': 'Note Number',
        'Dr./Cr. Date': 'Note date',
        'Type of note                (Dr/ Cr)': 'Note Type',
        'Place of supply': 'Place Of Supply',
        'RCM': 'Reverse Charge',
        'Invoice Type': 'Note Supply Type',
        'Dr./Cr. Value': 'Note Value',
        'GST%': 'Rate',
        'Taxable Value': 'Taxable Value',
        'CESS': 'Cess Amount'
    },
    'cdnur': {
        'Supply Type': 'UR Type',
        'Dr./ Cr. Note No.': 'Note Number',
        'Dr./ Cr. Note Date': 'Note date',
        'Type of note (Dr./ Cr.)': 'Note Type',
        'Place of supply': 'Place Of Supply',
        'Dr./Cr. Note Value': 'Note Value',
        'GST%': 'Rate',
        'Taxable Value': 'Taxable Value',
        'CESS': 'Cess Amount'
    },
    'adv_tax': {
        'Place Of Supply': 'Place Of Supply',
        'GST%': 'Rate',
        'Gross Advance Received': 'Gross Advance Received',
        'CESS': 'Cess Amount'
    },
    'adv_tax_adjusted': {
        'Place Of Supply': 'Place Of Supply',
        'GST%': 'Rate',
        'Gross Advance Adjusted': 'Gross Advance Adjusted',
        'CESS': 'Cess Amount'
    },
    'Docs_issued': {
        'Nature of Document': 'Nature of Document',
        'Sr. No. From': 'Sr.No.From',
        'Sr. No. To': 'Sr.No.To',
        'Total Number': 'Total Number',
        'Cancelled': 'Cancelled',
        'Net Issued': 'Net Issued' # Calculated field
    },
    'hsn': {
        'Type': 'Type', # Calculated
        'HSN': 'HSN',
        'Description': 'Description',
        'UQC': 'UQC',
        'Total Quantity': 'Total Quantity',
        'Total Value': 'Total Value',
        'Rate': 'Rate',
        'Total Taxable Value': 'Taxable Value',
        'IGST': 'Integrated Tax Amount',
        'CGST': 'Central Tax Amount',
        'SGST': 'State/UT Tax Amount',
        'CESS': 'Cess Amount'
    }
}

# ─── Helper Functions ─────────────────────────────────────────────────────────

def get_sheet_name_from_file(file_name: str) -> Optional[str]:
    upper = file_name.upper()
    if 'HSN' in upper: return 'hsn'
    if 'B2B' in upper: return 'b2b'
    if 'B2CL' in upper: return 'b2cl'
    if 'B2CS' in upper: return 'b2cs'
    if 'EXP' in upper: return 'export'
    if 'EXEMP' in upper: return 'Nil_exempt_NonGST'
    if 'CDNR' in upper and 'CDNUR' not in upper: return 'cdnr'
    if 'CDNUR' in upper: return 'cdnur'
    if 'ATADJ' in upper: return 'adv_tax_adjusted'
    if 'AT' in upper and 'ATADJ' not in upper: return 'adv_tax'
    if 'DOC' in upper: return 'Docs_issued'
    return None

def clean_place_of_supply(value: str) -> str:
    if not value: return ''
    # Remove leading digits and hyphen, e.g. "33-Tamil Nadu" -> "Tamil Nadu"
    return re.sub(r'^\d+-\s*', '', str(value)).strip()

def parse_date(value: str) -> Any:
    """Try to parse DD-MMM-YY to DD-MM-YYYY."""
    if not value: return value
    # Regex for DD-MMM-YY
    match = re.search(r'(\d{1,2})-([A-Za-z]{3})-(\d{2})', str(value))
    if match:
        day, month_str, year_suffix = match.groups()
        try:
            dt = datetime.strptime(f"{day}-{month_str}-{year_suffix}", "%d-%b-%y")
            return dt.strftime("%d-%m-%Y")
        except ValueError:
            pass
    return value

def process_docs_issued(rows: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    processed = []
    for row in rows:
        cleaned = {k: v.strip() for k, v in row.items()}
        total_number = float(cleaned.get('Total Number', 0) or 0)
        cancelled = float(cleaned.get('Cancelled', 0) or 0)
        
        # Handle variations in CSV headers
        nature = cleaned.get('Nature of Document') or cleaned.get('Type of Document') or ''
        sr_from = cleaned.get('Sr.No.From') or cleaned.get('Sr. No. From') or cleaned.get('Series From') or ''
        sr_to = cleaned.get('Sr.No.To') or cleaned.get('Sr. No. To') or cleaned.get('Series To') or ''

        processed.append({
            'Nature of Document': nature,
            'Sr.No.From': sr_from,
            'Sr.No.To': sr_to,
            'Total Number': total_number,
            'Cancelled': cancelled,
            'Net Issued': total_number - cancelled
        })
    return processed

def process_hsn_data(rows: List[Dict[str, str]], file_name: str) -> List[Dict[str, Any]]:
    is_b2b = 'B2B' in file_name.upper()
    is_b2c = 'B2C' in file_name.upper()
    
    processed = []
    for row in rows:
        cleaned = {k: v.strip() for k, v in row.items()}
        if is_b2b: cleaned['Type'] = 'B2B'
        elif is_b2c: cleaned['Type'] = 'B2C'
        
        if not cleaned.get('Rate'):
            cleaned['Rate'] = 0
            
        processed.append(cleaned)
    return processed

def parse_csv_content(content: str) -> List[Dict[str, str]]:
    f = io.StringIO(content)
    # Handle optional BOM
    if content.startswith('\ufeff'):
        f = io.StringIO(content[1:])
    reader = csv.DictReader(f)
    if not reader.fieldnames:
        return []
    return list(reader)

# ─── Excel Updating Functions ─────────────────────────────────────────────────

def update_exempt_sheet(ws, data: List[Dict[str, Any]], mapping_dict: Dict[str, str]):
    """
    Updates the Exempt sheet by matching 'Description' in Column A.
    mapping_dict: {Excel_Header: CSV_Header} -> Inverted from standard usage above.
    Wait, STANDARD_MAPPINGS uses Excel_Header: CSV_Header? No.
    Let's check: 'b2b': {'GSTIN/UIN': 'GSTIN/UIN of Recipient'} -> Key is Excel, Value is CSV?
    The JS code says:
    mapping = columnMappings[sheetName]
    ...
    for (const [csvCol, excelCol] of Object.entries(columnMapping))
    
    In JS:
    const standardMappings = { 'b2b': { 'GSTIN/UIN of Recipient': 'GSTIN/UIN' } }
    Key = CSV Header, Value = Excel Header.
    
    My Python STANDARD_MAPPINGS above definition:
    'b2b': { 'GSTIN/UIN': 'GSTIN/UIN of Recipient' } -> I inverted it inadvertently?
    JS: 'GSTIN/UIN of Recipient': 'GSTIN/UIN'
    Python above: 'GSTIN/UIN': 'GSTIN/UIN of Recipient' (Excel Header -> CSV Header)
    
    I should verify usage.
    append_data_to_sheet needs to map CSV Data -> Excel Column.
    Usage: value = row_data.get(csv_col, '')
    So we need CSV Header to lookup value.
    The mapping should be used as: excel_header -> csv_header or vice versa.
    
    Let's FIX STANDARD_MAPPINGS to match JS exactly: {CSV_Header: Excel_Header}
    """
    if not data: return

    # Get description map from template (row index -> description)
    # Assuming standard template format: Descriptions in Column A
    template_desc_map = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
        if row[0]:
            template_desc_map[str(row[0]).strip()] = row_idx

    # Get header map (column name -> col index)
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    for row_data in data:
        desc = str(row_data.get('Description', '')).strip()
        if desc in template_desc_map:
            row_idx = template_desc_map[desc]
            
            # Use mapping: CSV_Key -> Excel_Key
            for csv_key, excel_key in mapping_dict.items():
                if excel_key in headers:
                    col_idx = headers[excel_key]
                    value = row_data.get(csv_key, '')
                    
                    if value == '':
                        # Keep existing or set to 0? JS sets to 0 if 't'='n'
                        # We'll just write it
                        continue
                        
                    # Convert to float if possible
                    try:
                        clean_val = str(value).replace(',', '')
                        if clean_val and clean_val.replace('.','',1).isdigit():
                            ws.cell(row=row_idx, column=col_idx, value=float(clean_val))
                        else:
                            ws.cell(row=row_idx, column=col_idx, value=str(value))
                    except ValueError:
                        ws.cell(row=row_idx, column=col_idx, value=str(value))


def append_data_to_sheet(ws, data: List[Dict[str, Any]], mapping_dict: Dict[str, str]):
    """
    Appends data to the sheet based on column mapping.
    mapping_dict: {CSV_Header: Excel_Header}
    """
    if not data: return
    
    # 1. Map Headers (Excel Header -> Column Index)
    # Scan Row 1
    headers = {}
    for cell in ws[1]:
        if cell.value is not None:
            # Normalize whitespace
            headers[str(cell.value).strip()] = cell.column
            
    # 2. Find Start Row
    start_row = ws.max_row + 1
    # Ensure we aren't appending after empty rows
    for r in range(ws.max_row, 1, -1):
        # Check if row is empty
        is_empty = all(ws.cell(row=r, column=c).value is None for c in range(1, ws.max_column + 1))
        if not is_empty:
            start_row = r + 1
            break
            
    # 3. Write Data
    for row_data in data:
        for csv_key, excel_key in mapping_dict.items():
            if excel_key in headers:
                col_idx = headers[excel_key]
                val = row_data.get(csv_key, '')
                
                # Apply transformations
                if 'Place Of Supply' in excel_key:
                    val = clean_place_of_supply(val)
                elif 'Invoice Type' in excel_key:
                    val = str(val).replace(' B2B', '').replace(' B2C', '').strip()
                elif excel_key in ['RCM Applicable', 'Reverse Charge']:
                    if val == 'Y': val = 'Yes'
                    elif val == 'N': val = 'No'
                elif 'Date' in excel_key:
                    val = parse_date(val)
                
                # Convert numbers
                if isinstance(val, str) and val.replace('.','',1).isdigit():
                    try:
                        val = float(val)
                    except ValueError:
                        pass
                
                ws.cell(row=start_row, column=col_idx, value=val)
        
        start_row += 1


def generate_gstr1_excel(csv_files: List[Dict[str, str]]) -> tuple[bytes, int]:
    """
    Main function to process CSVs and generate GSTR1 Excel.
    """
    from pathlib import Path
    template_path = Path(__file__).resolve().parent / "templates" / "GSTR1_Template.xlsx"

    if template_path.exists():
        logger.info(f"Using template: {template_path}")
        wb = load_workbook(template_path)
    else:
        # Fallback if no template (should not happen based on user request)
        print("WARNING: Template not found at", template_path)
        wb = Workbook()

    sheets_processed = 0

    for csv_file in csv_files:
        file_name = csv_file["name"]
        content = csv_file["content"]
        
        sheet_name = get_sheet_name_from_file(file_name)
        if not sheet_name:
            # Try looser match?
            logger.warning(f"Skipping {file_name} - Unknown sheet type")
            continue
            
        try:
            raw_rows = parse_csv_content(content)
        except Exception as e:
            logger.error(f"Error parsing {file_name}: {e}")
            continue

        if not raw_rows:
            continue

        # Pre-process rows
        if sheet_name == 'Docs_issued':
            rows = process_docs_issued(raw_rows)
        elif sheet_name == 'hsn':
            rows = process_hsn_data(raw_rows, file_name)
        else:
            rows = [{k: v.strip() for k, v in row.items()} for row in raw_rows]

        # Get Worksheet
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet {sheet_name} missing in template. Creating new.")
            ws = wb.create_sheet(sheet_name)
            # Add headers? 
        else:
            ws = wb[sheet_name]

        # Get Mapping
        mapping = STANDARD_MAPPINGS.get(sheet_name, {})
        
        # Invert Mapping above if I used Excel:CSV.
        # Let's fix STANDARD_MAPPINGS to be CSV:Excel as per JS logic.
        # JS: 'GSTIN/UIN of Recipient': 'GSTIN/UIN' (CSV -> Excel)
        # My STANDARD_MAPPINGS below should match this structure.
        
        if sheet_name == 'Nil_exempt_NonGST':
            update_exempt_sheet(ws, rows, mapping)
        else:
            append_data_to_sheet(ws, rows, mapping)
            
        sheets_processed += 1
        logger.info(f"Processed {file_name} -> {sheet_name} ({len(rows)} rows)")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), sheets_processed

# ─── Corrected Mappings (CSV Header -> Excel Header) ──────────────────────────
# Re-declaring here to ensure correctness before main logic uses it.
STANDARD_MAPPINGS = {
    'b2b': {
        'GSTIN/UIN of Recipient': 'GSTIN/UIN',
        'Invoice Number': 'Invoice No',
        'Invoice date': 'Date of Invoice',
        'Invoice Value': 'Invoice Value',
        'Rate': 'GST%',
        'Taxable Value': 'Taxable Value',
        'Cess Amount': 'CESS',
        'Place Of Supply': 'Place Of Supply',
        'Reverse Charge': 'RCM Applicable',
        'Invoice Type': 'Invoice Type',
        'E-Commerce GSTIN': 'E-Commerce GSTIN'
    },
    'b2cl': {
        'Invoice Number': 'Invoice No',
        'Invoice date': 'Date of Invoice',
        'Invoice Value': 'Invoice Value',
        'Place Of Supply': 'Place Of Supply',
        'Rate': 'GST%',
        'Taxable Value': 'Taxable Value',
        'Cess Amount': 'CESS',
        'E-Commerce GSTIN': 'E-Commerce GSTIN'
    },
    'b2cs': {
        'Type': 'Type',
        'Place Of Supply': 'Place Of Supply',
        'Rate': 'GST%',
        'Taxable Value': 'Taxable Value',
        'Cess Amount': 'CESS',
        'E-Commerce GSTIN': 'E-Commerce GSTIN'
    },
    'export': {
        'Export Type': 'Export Type',
        'Invoice Number': 'Invoice No',
        'Invoice date': 'Date of Invoice',
        'Invoice Value': 'Invoice Value',
        'Port Code': 'Port Code',
        'Shipping Bill Number': 'Shipping Bill No',
        'Shipping Bill Date': 'Shipping Bill Date',
        'Rate': 'GST%',
        'Taxable Value': 'Taxable Value'
    },
    'Nil_exempt_NonGST': {
        'Description': 'Description',
        'Nil Rated Supplies': 'Nil Rated Supplies',
        'Exempted (other than nil rated/non GST supply)': 'Exempted(other than nil rated/non GST supply)',
        'Non-GST supplies': 'Non-GST Supplies'
    },
    'cdnr': {
        'GSTIN/UIN of Recipient': 'GSTIN/UIN',
        'Note Number': 'Dr./ Cr. No.',
        'Note Date': 'Dr./Cr. Date',
        'Note Type': 'Type of note                (Dr/ Cr)',
        'Place Of Supply': 'Place of supply',
        'Reverse Charge': 'RCM',
        'Note Supply Type': 'Invoice Type',
        'Note Value': 'Dr./Cr. Value',
        'Rate': 'GST%',
        'Taxable Value': 'Taxable Value',
        'Cess Amount': 'CESS'
    },
    'cdnur': {
        'UR Type': 'Supply Type',
        'Note/Refund Voucher Number': 'Dr./ Cr. Note No.',
        'Note/Refund Voucher date': 'Dr./ Cr. Note Date',
        'Document Type': 'Type of note (Dr./ Cr.)',
        'Place Of Supply': 'Place of supply',
        'Note/Refund Voucher Value': 'Dr./Cr. Note Value',
        'Rate': 'GST%',
        'Taxable Value': 'Taxable Value',
        'Cess Amount': 'CESS'
    },
    'adv_tax': {
        'Place Of Supply': 'Place Of Supply',
        'Rate': 'GST%',
        'Gross Advance Received': 'Gross Advance Received',
        'Cess Amount': 'CESS'
    },
    'adv_tax_adjusted': {
        'Place Of Supply': 'Place Of Supply',
        'Rate': 'GST%',
        'Gross Advance Adjusted': 'Gross Advance Adjusted',
        'Cess Amount': 'CESS'
    },
    'Docs_issued': {
        'Nature of Document': 'Nature of Document',
        'Sr.No.From': 'Sr. No. From',
        'Sr.No.To': 'Sr. No. To',
        'Total Number': 'Total Number',
        'Cancelled': 'Cancelled',
        'Net Issued': 'Net Issued'
    },
    'hsn': {
        'Type': 'Type',
        'HSN': 'HSN',
        'Description': 'Description',
        'UQC': 'UQC',
        'Total Quantity': 'Total Quantity',
        'Total Value': 'Total Value',
        'Rate': 'Rate',
        'Taxable Value': 'Total Taxable Value',
        'Integrated Tax Amount': 'IGST',
        'Central Tax Amount': 'CGST',
        'State/UT Tax Amount': 'SGST',
        'Cess Amount': 'CESS'
    }
}
