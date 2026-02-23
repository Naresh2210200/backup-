"""
GSTIN Verification Router
==========================
POST /verification/run
  - Downloads the GSTR1 Excel from R2
  - Reads B2B GSTINs, validates, moves invalids to B2C grouped by rate.
  - Validates HSN codes and reconciles B2B/B2C HSN totals to match B2B->B2C shifts.
  - Uploads corrected Excel + error report to R2
  - Returns run summary & dashboard totals
"""
import io
import re
import time
import logging
from typing import List, Optional, Dict, Any
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from collections import defaultdict
import difflib

import storage

logger = logging.getLogger(__name__)
router = APIRouter()

GSTIN_REGEX = re.compile(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$')

STANDARD_HSN_MASTER = {
    '1001': 'Wheat and meslin',
    '1002': 'Rye',
    '998412': 'Telecommunication services',
    '998599': 'Other support services',
    '847130': 'Portable digital ADP machines',
    '851712': 'Telephones for cellular networks',
    '998311': 'Management consulting',
    '998312': 'Business consulting',
    '9983': 'Professional Services',
    '8471': 'Computers and tech',
    '99': 'Service generic'
}

def fuzzy_match_hsn(hsn_code: str) -> Optional[str]:
    hsn_str = str(hsn_code).strip().split('.')[0]
    if hsn_str in STANDARD_HSN_MASTER:
        return hsn_str
    matches = difflib.get_close_matches(hsn_str, STANDARD_HSN_MASTER.keys(), n=1, cutoff=0.6)
    if matches:
        return matches[0]
    return hsn_str # default to itself if no match found but we log it

def clean_val(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        return float(str(v).replace(',','').strip())
    except:
        return 0.0

class VerificationRequest(BaseModel):
    storage_key: str
    ca_code: str
    customer_id: Optional[str] = None
    financial_year: Optional[str] = None
    month: Optional[str] = None

class VerificationSummary(BaseModel):
    run_id: str
    total_checked: int
    total_invalid: int
    total_moved_to_b2cs: int
    corrected_key: Optional[str] = None
    error_report_key: Optional[str] = None
    dashboard_data: Dict[str, Any] = {}
    status: str = "completed"

@router.post("/run", response_model=VerificationSummary)
async def run_verification(payload: VerificationRequest):
    import openpyxl
    from openpyxl import load_workbook, Workbook

    logger.info(f"Verification started for: {payload.storage_key}")

    # 1. Read file
    try:
        excel_bytes = storage.read_bytes(payload.storage_key)
    except Exception as read_err:
        logger.error(f"Failed local read_bytes: {str(read_err)}")
        try:
            import boto3
            from botocore.config import Config
            from config import settings
            endpoint = settings.R2_ENDPOINT_URL or f"https://{settings.R2_ACCOUNT_ID}.r2.cloudflarestorage.com"
            s3 = boto3.client(
                's3', endpoint_url=endpoint,
                aws_access_key_id=settings.R2_ACCESS_KEY_ID,
                aws_secret_access_key=settings.R2_SECRET_ACCESS_KEY,
                config=Config(signature_version='s3v4'), region_name='auto'
            )
            resp = s3.get_object(Bucket=settings.R2_BUCKET_NAME, Key=payload.storage_key)
            excel_bytes = resp['Body'].read()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Cannot read file: Invalid endpoint. Please set your R2_ACCOUNT_ID in .env or provide a valid endpoint. Internal Err: {e}")

    wb = load_workbook(io.BytesIO(excel_bytes))
    
    b2b_sheet = wb['b2b'] if 'b2b' in wb.sheetnames else None
    b2cs_sheet = wb['b2cs'] if 'b2cs' in wb.sheetnames else wb.create_sheet('b2cs')
    hsn_sheet = wb['hsn'] if 'hsn' in wb.sheetnames else None

    # Error Tracking
    errors = []
    total_checked = 0
    total_moved = 0
    
    b2cs_additions = defaultdict(lambda: {'Taxable Value': 0.0, 'Cess': 0.0}) # grouped by (POS, Rate)
    hsn_b2b_deductions = defaultdict(float) # rate -> taxable value shifted from B2B to B2C

    # --- MODULE 1: B2B VERIFICATION ---
    if b2b_sheet:
        headers = [str(cell.value).strip() if cell.value else '' for cell in b2b_sheet[1]]
        
        gstin_col = next((i for i, h in enumerate(headers) if 'GSTIN' in h.upper()), None)
        party_col = next((i for i, h in enumerate(headers) if 'NAME' in h.upper() or 'RECEIVER' in h.upper()), gstin_col) # Fallback to GSTIN if no name
        tax_col = next((i for i, h in enumerate(headers) if 'TAXABLE VALUE' in h.upper()), None)
        rate_col = next((i for i, h in enumerate(headers) if 'RATE' in h.upper() or 'GST%' in h.upper()), None)
        pos_col = next((i for i, h in enumerate(headers) if 'PLACE OF SUPPLY' in h.upper()), None)
        cess_col = next((i for i, h in enumerate(headers) if 'CESS' in h.upper()), None)

        if gstin_col is not None and tax_col is not None and rate_col is not None:
            rows_to_delete = []
            
            for row_idx, row in enumerate(b2b_sheet.iter_rows(min_row=2, values_only=False), start=2):
                gstin_val = row[gstin_col].value
                if not gstin_val: continue
                
                gstin = str(gstin_val).strip().upper()
                total_checked += 1

                is_valid = bool(GSTIN_REGEX.match(gstin))
                if not is_valid: # Fake unregistered / inactive checks can be added here
                    total_moved += 1
                    
                    party_name = str(row[party_col].value) if party_col is not None else "Unknown"
                    tax_val = clean_val(row[tax_col].value)
                    rate = clean_val(row[rate_col].value)
                    pos = str(row[pos_col].value) if pos_col is not None else "Unknown POS"
                    cess = clean_val(row[cess_col].value) if cess_col is not None else 0.0

                    errors.append({
                        "GSTIN": gstin,
                        "Party Name": party_name,
                        "Error Type": "Invalid/Unregistered GSTIN",
                        "Taxable Value": tax_val,
                        "Rate": rate,
                        "Action": "Moved to B2C"
                    })
                    
                    # Group by B2C rules (Place of Supply, Rate)
                    b2cs_additions[(pos, rate)]['Taxable Value'] += tax_val
                    b2cs_additions[(pos, rate)]['Cess'] += cess
                    
                    # Tell HSN reconciler that this rate's B2B is transitioning to B2C
                    hsn_b2b_deductions[rate] += tax_val
                    
                    rows_to_delete.append(row_idx)

            # Move invalid invoices out of B2B completely
            for row_idx in sorted(rows_to_delete, reverse=True):
                b2b_sheet.delete_rows(row_idx, 1)

    # Move added B2CS entries into the actual B2CS sheet
    if b2cs_additions:
        # Check B2CS headers
        b2cs_headers = [str(cell.value).strip() if cell.value else '' for cell in b2cs_sheet[1]]
        if not b2cs_headers or not any(b2cs_headers):
            b2cs_headers = ['Type', 'Place Of Supply', 'Rate', 'Taxable Value', 'Cess Amount', 'E-Commerce GSTIN']
            for col_idx, h in enumerate(b2cs_headers, 1):
                b2cs_sheet.cell(row=1, column=col_idx, value=h)
                
        type_c = b2cs_headers.index('Type') + 1 if 'Type' in b2cs_headers else 1
        pos_c = b2cs_headers.index('Place Of Supply') + 1 if 'Place Of Supply' in b2cs_headers else 2
        rate_c = next((i+1 for i, h in enumerate(b2cs_headers) if 'RATE' in h.upper() or 'GST%' in h.upper()), 3)
        tax_c = next((i+1 for i, h in enumerate(b2cs_headers) if 'TAXABLE VALUE' in h.upper()), 4)
        cess_c = next((i+1 for i, h in enumerate(b2cs_headers) if 'CESS' in h.upper()), 5)

        start_row = b2cs_sheet.max_row + 1
        for (pos, rate), amounts in b2cs_additions.items():
            b2cs_sheet.cell(row=start_row, column=type_c, value='OE')  # Other than e-commerce
            b2cs_sheet.cell(row=start_row, column=pos_c, value=pos)
            b2cs_sheet.cell(row=start_row, column=rate_c, value=rate)
            b2cs_sheet.cell(row=start_row, column=tax_c, value=amounts['Taxable Value'])
            b2cs_sheet.cell(row=start_row, column=cess_c, value=amounts['Cess'])
            start_row += 1

    # --- MODULE 2 & RECONCILIATION: HSN VERIFICATION ---
    if hsn_sheet and hsn_sheet.max_row > 1:
        headers = [str(cell.value).strip() if cell.value else '' for cell in hsn_sheet[1]]
        type_col = next((i for i, h in enumerate(headers) if 'TYPE' in h.upper()), None)
        hsn_col = next((i for i, h in enumerate(headers) if 'HSN' in h.upper()), None)
        rate_col = next((i for i, h in enumerate(headers) if 'RATE' in h.upper()), None)
        tax_col = next((i for i, h in enumerate(headers) if 'TAXABLE VALUE' in h.upper() and 'TOTAL' not in h.upper()), None)
        
        if not tax_col:  # fallback
            tax_col = next((i for i, h in enumerate(headers) if 'TAXABLE' in h.upper()), None)

        if hsn_col is not None and tax_col is not None and rate_col is not None:
            # 1. Fix fuzzy HSN matched codes
            for row in hsn_sheet.iter_rows(min_row=2, values_only=False):
                h_val = row[hsn_col].value
                if h_val:
                    correct_hsn = fuzzy_match_hsn(str(h_val))
                    if correct_hsn and correct_hsn != str(h_val).strip():
                        # We log this correction implicitly or explicitly
                        row[hsn_col].value = correct_hsn
            
            # 2. Difference Fixing: Re-apportion B2B -> B2C Taxable Values in HSN
            if type_col is not None and hsn_b2b_deductions:
                for target_rate, amt_to_shift in hsn_b2b_deductions.items():
                    if amt_to_shift <= 0: continue
                    
                    b2b_target_rows = []
                    b2c_target_rows = []
                    
                    # Find all HSN rows matching this rate
                    for row_idx, row in enumerate(hsn_sheet.iter_rows(min_row=2, values_only=False), start=2):
                        item_type = str(row[type_col].value).strip().upper() if row[type_col].value else ''
                        item_rate = clean_val(row[rate_col].value)
                        
                        if item_rate == target_rate:
                            if item_type == 'B2B':
                                b2b_target_rows.append(row)
                            elif item_type == 'B2C':
                                b2c_target_rows.append(row)
                    
                    # Deduct from B2B HSNs (greedily)
                    remaining_to_shift = amt_to_shift
                    for rb in b2b_target_rows:
                        curr_tax = clean_val(rb[tax_col].value)
                        deduct_amt = min(curr_tax, remaining_to_shift)
                        rb[tax_col].value = curr_tax - deduct_amt
                        
                        # Apply corresponding additions to B2C HSNs if exist, else we just append a B2C row
                        # Here, for simplicity, we add it to the first B2C row with same rate, or duplicate this HSN row as B2C
                        matched_b2c = next((rc for rc in b2c_target_rows if str(rc[hsn_col].value) == str(rb[hsn_col].value)), None)
                        if matched_b2c:
                            matched_b2c[tax_col].value = clean_val(matched_b2c[tax_col].value) + deduct_amt
                        else:
                            # Add a new B2C row for this HSN
                            max_r = hsn_sheet.max_row + 1
                            for c_idx, cell in enumerate(rb):
                                r_c = hsn_sheet.cell(row=max_r, column=c_idx+1)
                                r_c.value = cell.value 
                            hsn_sheet.cell(row=max_r, column=type_col+1).value = 'B2C'
                            hsn_sheet.cell(row=max_r, column=tax_col+1).value = deduct_amt
                        
                        remaining_to_shift -= deduct_amt
                        if remaining_to_shift <= 0:
                            break

    # 4. Generate Error Report CSV
    # Instead of pulling pandas, manually compose CSV string 
    if errors:
        headers = list(errors[0].keys())
        error_csv = ",".join(headers) + "\n"
        for row in errors:
            error_csv += ",".join(str(row[k]) for k in headers) + "\n"
    else:
        error_csv = "GSTIN,Party Name,Error Type,Taxable Value,Rate,Action\n"

    # 5. Build Dashboard Summary (Compute new totals across sheets)
    dash_data = {
        "b2b_taxable": 0.0, "b2b_igst": 0.0, "b2b_cgst": 0.0, "b2b_sgst": 0.0, "b2b_cess": 0.0,
        "b2c_taxable": 0.0, "b2c_igst": 0.0, "b2c_cgst": 0.0, "b2c_sgst": 0.0, "b2c_cess": 0.0,
        "hsn_taxable": 0.0, "hsn_igst": 0.0, "hsn_cgst": 0.0, "hsn_sgst": 0.0, "hsn_cess": 0.0,
        "cdnr_taxable": 0.0, "cdnr_igst": 0.0, "cdnr_cgst": 0.0, "cdnr_sgst": 0.0, "cdnr_cess": 0.0,
    }
    
    implicit_taxes_by_pos = defaultdict(float)
    implicit_cats_by_pos = defaultdict(lambda: {'b2b': 0.0, 'b2c': 0.0, 'cdnr': 0.0})
    
    def sum_sheet(s_name, prefix):
        # FIX: wb.get_sheet_by_name() is deprecated in openpyxl — use wb[s_name] instead
        if s_name not in wb.sheetnames:
            return
        ws = wb[s_name]
        if not ws or ws.max_row <= 1:
            return
        head = [str(cell.value).strip().upper() if cell.value else '' for cell in ws[1]]
        # Taxable value — exclude 'TOTAL VALUE' but include 'TAXABLE VALUE'
        tax_c = next((i for i, h in enumerate(head) if 'TAXABLE' in h), None)
        if tax_c is None:
            # Fallback: use 'TOTAL VALUE' for HSN-like sheets
            tax_c = next((i for i, h in enumerate(head) if 'TOTAL VALUE' in h), None)
        # Tax columns — broadened to match 'CGST AMOUNT', 'IGST AMT', etc.
        i_c = next((i for i, h in enumerate(head) if 'IGST' in h or 'INTEGRATED' in h), None)
        c_c = next((i for i, h in enumerate(head) if 'CGST' in h or 'CENTRAL' in h), None)
        s_c = next((i for i, h in enumerate(head) if ('SGST' in h or 'STATE' in h) and 'CESS' not in h), None)
        cess_c = next((i for i, h in enumerate(head) if 'CESS' in h), None)
        
        # Heuristic setup for implicit calculations:
        rate_c = next((i for i, h in enumerate(head) if 'RATE' in h or 'GST%' in h), None)
        pos_c = next((i for i, h in enumerate(head) if 'PLACE OF SUPPLY' in h), None)
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            tax_v = clean_val(row[tax_c]) if tax_c is not None else 0.0
            dash_data[f"{prefix}_taxable"] += tax_v
            
            if cess_c is not None: dash_data[f"{prefix}_cess"] += clean_val(row[cess_c])
            
            # Explicit taxes present 
            if c_c is not None or i_c is not None:
                if i_c is not None: dash_data[f"{prefix}_igst"] += clean_val(row[i_c])
                if c_c is not None: dash_data[f"{prefix}_cgst"] += clean_val(row[c_c])
                if s_c is not None: dash_data[f"{prefix}_sgst"] += clean_val(row[s_c])
            else:
                # Implicit taxes - gather dynamically to be apportioned after HSN is parsed
                rate_v = clean_val(row[rate_c]) if rate_c is not None else 0.0
                row_tax = tax_v * (rate_v / 100.0)
                if pos_c is not None and row[pos_c]:
                    # Normalize "29-Karnataka" to "KARNATAKA" to perfectly match heuristic state keys
                    pos_str = re.sub(r'^\d+[-\s]+', '', str(row[pos_c])).strip().upper()
                else:
                    pos_str = 'UNKNOWN'
                    
                implicit_taxes_by_pos[pos_str] += row_tax
                implicit_cats_by_pos[pos_str][prefix] += row_tax

    sum_sheet('b2b', 'b2b')
    sum_sheet('b2cs', 'b2c')
    sum_sheet('b2cl', 'b2c') # Also aggregate B2CL into B2C total
    sum_sheet('hsn', 'hsn')
    sum_sheet('cdnr', 'cdnr')
    sum_sheet('cdnur', 'cdnr')

    # Apply Heuristic Distribution for Missing Explicit Taxes:
    if implicit_taxes_by_pos:
        hsn_cgst = dash_data.get('hsn_cgst', 0.0)
        
        # 1. Identify which Place of Supply is the "Home State" by minimising delta error with HSN CGST
        best_state, min_diff = None, float('inf')
        for candidate_state, candidate_gross_tax in implicit_taxes_by_pos.items():
            diff = abs((candidate_gross_tax / 2) - hsn_cgst)
            if diff < min_diff:
                min_diff = diff
                best_state = candidate_state
                
        # 2. Assign Implicit tax proportions
        for pos, cats in implicit_cats_by_pos.items():
            is_home = (pos == best_state)
            for prefix, gross_tax in cats.items():
                if is_home:
                    dash_data[f"{prefix}_cgst"] += gross_tax / 2
                    dash_data[f"{prefix}_sgst"] += gross_tax / 2
                else:
                    dash_data[f"{prefix}_igst"] += gross_tax

    # Total Tax = IGST + CGST + SGST
    for pfx in ['b2b', 'b2c', 'hsn', 'cdnr']:
        dash_data[f"{pfx}_total_tax"] = dash_data[f"{pfx}_igst"] + dash_data[f"{pfx}_cgst"] + dash_data[f"{pfx}_sgst"]

    timestamp = int(time.time())
    base_path = f"outputs/{payload.ca_code}/{payload.customer_id or 'unknown'}"
    corrected_key = f"{base_path}/corrected_{timestamp}.xlsx"
    error_key = f"{base_path}/error_report_{timestamp}.csv"

    corrected_buf = io.BytesIO()
    wb.save(corrected_buf)
    storage.save_file(corrected_key, corrected_buf.getvalue(),
                      'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    storage.save_file(error_key, error_csv.encode('utf-8'), 'text/csv')

    import uuid
    run_id = str(uuid.uuid4())

    return VerificationSummary(
        run_id=run_id,
        total_checked=total_checked,
        total_invalid=len(errors),
        total_moved_to_b2cs=total_moved,
        corrected_key=corrected_key,
        error_report_key=error_key,
        dashboard_data=dash_data
    )
